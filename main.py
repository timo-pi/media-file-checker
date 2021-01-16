from openpyxl import Workbook
import datetime

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Strg+F8 to toggle the breakpoint.

    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = name

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted

    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")


if __name__ == '__main__':
    print_hi('Timo')
